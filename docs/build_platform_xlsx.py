"""platform 설계서 V3 빌드 — admin V4 형식 (5컬럼) + 자연어 톤."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = 'C:/users/campus3S026/LS/lifesync360-platform 설계서 V3.xlsx'

# ── admin V4 스타일 상수 ──────────────────────────────────────────
FONT       = Font(name='맑은 고딕', size=12)
FONT_BOLD  = Font(name='맑은 고딕', size=12, bold=True)
FILL_TITLE = PatternFill('solid', fgColor='D9E1F2')  # 섹션 제목 (옅은 파랑)
FILL_HEAD  = PatternFill('solid', fgColor='FFF2CC')  # 컬럼 헤더 (옅은 노랑)
ALIGN_C    = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_L    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
BORDER     = Border(*[Side(style='thin', color='BFBFBF')] * 4)
COL_WIDTHS = {'A': 24, 'B': 36, 'C': 32, 'D': 36, 'E': 68, 'F': 20}
HEADERS    = ['기능', '설명', '데이터 소스', '테이블 / 객체', '컬럼 / 비고']
HEADERS_API = ['API', '설명', '구현 위치', '엔드포인트', '데이터 소스']


# ── 일반 섹션 (시트별) ────────────────────────────────────────────
SHEETS = {
    '인증 · 로그인': [
        ('로그인 · 토큰 발급 (1번 영역 : 시연 진입)', [
            ('로그인 화면',
             '이메일·비밀번호를 입력받아 백엔드에 로그인 요청',
             '프론트엔드 SSR',
             'templates/login.html',
             '비밀번호 입력 → 로그인 버튼 → /api/login 호출'),
            ('계정 검증',
             '입력한 이메일·비밀번호가 5 등급 데모 계정과 일치하는지 확인',
             'test_login_credentials.csv (시연용 사전 발급 계정)',
             'login_email + sha256_hex',
             '5 등급 × 1 명 — VIP / GOLD / SILVER / BASIC / CARE (예: user0000924@lifesync.com · DemoVIP01!). 비밀번호는 SHA-256 해시 비교'),
            ('토큰 발급 방식',
             '로그인 검증 통과 시 플랫폼이 secret 으로 HS256 토큰 직접 발급. 응답 본문에 token 반환',
             'AWS Systems Manager (Parameter Store)',
             '/lifesync360/jwt-secret (SecureString)',
             'jwt.encode({sub, gid, exp}, secret, HS256) — secret 은 발급/검증 양쪽에 동일'),
            ('토큰 검증 시크릿',
             '클라이언트가 들고 오는 토큰을 같은 secret 으로 검증 (HS256 대칭키)',
             'AWS Systems Manager (Parameter Store)',
             '/lifesync360/jwt-secret (SecureString)',
             '첫 호출 1회만 SSM 조회 후 메모리 캐시. ECS 배포 시 환경변수로 자동 주입'),
        ]),
        ('본인 프로필 조회 (2번 영역 : 헤더 / 마이페이지)', [
            ('이름 · 등급',
             '헤더 상단에 표시할 본인 이름과 등급',
             '온프레미스 + DynamoDB',
             'customer_pii_secure + lifesync_customer_result',
             '이름은 마스킹(예: 김*수), 등급은 VIP/GOLD/SILVER/BASIC/CARE'),
            ('인구통계',
             '성별·연령대·지역·소득·자산 등급 등',
             '온프레미스',
             'customer_360_profile',
             '성별 M/F, 연령대 20~60대+, 지역 SEOUL/GYEONGGI 등'),
            ('가입 · 마스터 정보',
             '가입일·최근 로그인 시각·고객 상태·고객 타입',
             '온프레미스',
             'master_customer',
             '활성(ACTIVE), 개인(INDIVIDUAL), 최초 가입일·최근 로그인'),
            ('동의 현황',
             '본인이 동의한 데이터 활용 도메인 목록',
             '온프레미스',
             'consent_master',
             '8개 도메인 (은행·카드·보험·증권·다이렉트보험·헬스케어·병원·웨어러블)'),
        ]),
        ('보호 API 진입 (3번 영역 : 모든 /api/* 호출 공통)', [
            ('JWT 검증 미들웨어',
             '클라이언트 토큰이 유효한지 확인 후 본인 정보를 함수에 전달',
             'app.py · @require_jwt',
             '-',
             '토큰 만료/위조 시 401 응답, 정상이면 sub(ls_user_id) / gid(global_id) 추출'),
            ('토큰 저장 위치',
             '로그인 후 받은 토큰을 브라우저에 보관',
             '브라우저 localStorage',
             'ls_token',
             '모든 API 호출 시 Authorization 헤더에 자동으로 첨부'),
        ]),
    ],

    '홈 · 점수 · 캠페인': [
        ('홈 페이지 (1번 영역 : 메인 진입 / 3 탭)', [
            ('홈 라우트',
             '로그인 안 한 경우 로그인 페이지로 이동, 그렇지 않으면 메인 화면 렌더',
             '프론트엔드 SSR',
             'templates/index.html',
             '메인은 3 탭으로 구성 (홈 / 추천 / MY)'),
            ('하단 탭 네비게이션',
             '홈·추천·MY 탭 전환 — 점수+캠페인 / 추천 top10 / 신청 내역',
             '클라이언트 JS',
             '.nav-btn / .tab-content',
             '탭 클릭 시 분석용 이벤트 자동 발사'),
        ]),
        ('점수 카드 (2번 영역 : 홈 탭 상단)', [
            ('종합 · 건강 점수',
             '내 점수를 원형 게이지로 표시. 토글로 종합 / 건강 전환',
             'DynamoDB',
             'lifesync_customer_result',
             '0~100 점수, 매일 새벽 ETL 로 갱신'),
        ]),
        ('등급별 캠페인 배너 (3번 영역 : 홈 탭 하단)', [
            ('캠페인 리스트',
             '내 등급에 맞는 캠페인 최대 5건 (시작일 최신순)',
             'Aurora MySQL',
             'campaign_master',
             '활성 상태 + 만료 안 된 캠페인만 표시'),
            ('등급 분기',
             'DynamoDB 의 동적 등급 기준으로 캠페인 매칭',
             'DynamoDB',
             'lifesync_customer_result.dynamic_grade',
             'VIP / GOLD / SILVER / BASIC / CARE'),
            ('CTA 클릭 이벤트',
             '배너 버튼 클릭 시 분석 데이터 적재 (어떤 캠페인이 인기인지 측정)',
             'Aurora MySQL',
             'customer_dashboard_log',
             '/api/event 호출 — 화면 이동 없음 (조용히 적재)'),
        ]),
    ],

    '추천 · 상품': [
        ('추천 리스트 — Aurora top 10 (1번 영역 : 추천 탭)', [
            ('① 점수 · 등급 조회',
             '추천 정렬에 필요한 본인 점수·등급·NBA 가져오기',
             'DynamoDB',
             'lifesync_customer_result',
             '등급·종합점수·건강점수·VIP 확률·다음 행동(NBA, 내부 사용)'),
            ('② 캐시 확인',
             '6시간 이내 추천 결과가 있으면 그대로 재사용',
             'AWS ElastiCache (Redis)',
             'rec:{global_id}',
             '캐시 히트 시 SQL 단계 생략 → 응답 빠름'),
            ('③ 룰 매칭',
             '등급·점수·건강조건에 맞는 추천 룰 찾기',
             'Aurora MySQL',
             'recommend_rule',
             'NBA 매칭되는 룰을 최상단으로 정렬 (내부 가중치)'),
            ('④ 교차 추천 보강',
             '주 카테고리와 어울리는 다른 카테고리 3개 추가',
             'Aurora MySQL',
             'cross_sell_rule',
             '예: 은행 적금 → 카드·보험 추천 함께'),
            ('⑤ 상품 매칭',
             '카테고리별로 우선순위 상위 2개씩, 최대 10개 추출',
             'Aurora MySQL',
             'product_master + company + category',
             '상품·계열사·카테고리 정보 JOIN, 우선순위 기준 정렬'),
            ('⑥ 이력 적재 · 캐시 갱신',
             '추천 결과를 이력 테이블에 기록하고 Redis 6시간 캐시',
             'Aurora MySQL + Redis',
             'customer_recommend_history',
             '나중에 클릭율·구매율 분석 데이터로 활용'),
        ]),
        ('상품 상세 (2번 영역 : 추천 카드 클릭 후)', [
            ('상품 정보',
             '상품 이름·설명·대상등급·위험도·옵션 등 상세 정보',
             'Aurora MySQL',
             'product_master + product_option',
             '회사 / 카테고리 정보까지 JOIN 으로 함께 조회'),
            ('상품 조회 이벤트',
             '상세 페이지 진입 시 자동으로 분석 데이터 적재',
             'Aurora MySQL',
             'customer_dashboard_log',
             '/api/event 호출 — 어떤 상품이 많이 보이는지 측정'),
        ]),
        ('상품 신청 (3번 영역 : 상품 상세에서 진입)', [
            ('신청 폼 페이지',
             '신청자 정보와 상품 정보를 함께 보여주는 폼 화면',
             '프론트엔드 SSR',
             'templates/apply.html',
             '추가 정보 입력 없이 단순 확인 후 신청 가능'),
            ('신청 처리',
             '신청서 접수 + 추천 이력 갱신 + 분석 로그 동시 기록',
             'Aurora MySQL',
             'customer_product_application',
             '신청번호 자동 생성 (APP-{시각}-{사용자식별 6자리}), 상태 RECEIVED'),
            ('관련 이력 갱신',
             '추천에서 들어온 상품인 경우 구매 플래그 자동 업데이트',
             'Aurora MySQL',
             'customer_recommend_history',
             '추천 클릭 → 신청 전환율 분석에 활용'),
        ]),
    ],

    '이벤트 · 신청 내역': [
        ('이벤트 적재 — 사용자 행동 추적 (1번 영역 : 분석 데이터)', [
            ('클릭 / 페이지뷰',
             '추천 클릭·상품 조회·배너 클릭 등 사용자 행동을 백엔드로 전송',
             'Aurora MySQL',
             'customer_dashboard_log',
             '응답을 기다리지 않고 백그라운드로 전송 (화면 영향 없음)'),
            ('추천 클릭 / 구매 갱신',
             '특정 이벤트는 추천 이력에 클릭·구매 플래그까지 갱신',
             'Aurora MySQL',
             'customer_recommend_history',
             '추천 클릭 → clicked_flag, 신청 완료 → purchased_flag'),
            ('이벤트 종류 (8개)',
             '추천 클릭 · 상품 조회 · 신청 시작 · 신청 완료 · 배너 클릭 · 탭 클릭 등',
             '클라이언트 JS',
             '-',
             '각 이벤트마다 페이지 타입과 클릭 종류가 자동 분기'),
        ]),
        ('내 신청 내역 (2번 영역 : MY 탭)', [
            ('신청 리스트',
             '내가 신청한 상품을 최근 50건까지 시간순으로 표시',
             'Aurora MySQL',
             'customer_product_application',
             '상품 / 회사 / 카테고리 정보 함께 JOIN'),
            ('진행 상태',
             '신청서가 어느 단계까지 갔는지 표시',
             'Aurora MySQL',
             'customer_product_application.status',
             '접수 → 심사중 → 승인 / 거절 / 취소'),
        ]),
    ],

    '동의 · 설정': [
        ('동의 입력 화면 (1번 영역 : 별도 진입)', [
            ('도메인 8개',
             '데이터 활용 동의 항목 — 은행·카드·보험·증권·다이렉트보험·헬스케어·병원·웨어러블',
             'app.py 메모리',
             'CONSENTS 정의',
             '각 항목마다 아이콘·라벨·설명·수집 범위 표시'),
            ('동의 저장',
             '체크박스로 선택한 항목을 백엔드에 저장',
             '온프레미스 PrivateAPI',
             'consent_master (운영)',
             '운영 시 Lambda 경유로 사내 DB 에 반영'),
        ]),
        ('설정 화면 (2번 영역 : 마이페이지 진입)', [
            ('등급 혜택',
             '내 등급에서 받을 수 있는 혜택 리스트',
             'app.py 메모리',
             'GRADE_BENEFITS 정의',
             '5등급 × 혜택 2~5개 (예: VIP — 보험료 15% 할인, PB 매니저 등)'),
            ('동의 관리',
             '동의 화면과 동일한 8개 도메인 — 여기서도 변경 가능',
             'app.py 메모리 / 온프레미스',
             'CONSENTS + consent_master',
             '설정 화면에서 변경해도 /api/consent 로 동일하게 저장'),
            ('진입 경로',
             '메인 헤더의 사용자 이름 클릭 시 이동',
             '프론트엔드 SSR',
             'templates/settings.html',
             '메인 어디서든 헤더로 접근 가능'),
        ]),
    ],

    '인프라 · 운영': [
        ('환경 변수 (1번 영역 : ECS 컨테이너 설정)', [
            ('인증',
             'HS256 토큰 발급·검증 시크릿 (대칭키)',
             'AWS Systems Manager',
             '/lifesync360/jwt-secret',
             'SecureString — ECS task secrets 로 자동 주입'),
            ('Service-DB 연결',
             'Aurora MySQL 호스트·계정·비밀번호·DB명',
             'AWS Secrets Manager / SSM',
             'lifesync360 DB',
             '커넥션은 요청마다 새로 열고 닫음 (pymysql)'),
            ('Redis 추천 캐시',
             '추천 결과 캐시 (TTL 6시간)',
             'AWS ElastiCache (Redis)',
             'rec:{global_id}',
             'VPC 보안그룹으로 통신 — 외부 노출 X'),
            ('DynamoDB',
             '점수·등급·NBA 등 ML 산출물 저장 테이블',
             'AWS DynamoDB',
             'lifesync_customer_result',
             '매일 새벽 GCP Vertex AI ETL 로 적재됨'),
            ('PrivateAPI Lambda',
             '온프레미스 데이터 (PII / 동의 / 사용자) 접근 경로',
             'AWS Lambda',
             'lifesync-onprem-customer-query',
             'VPN 경유 → Control Node → 사내 MySQL'),
        ]),
        ('헬스체크 (2번 영역 : ALB 트래픽 라우팅 기준)', [
            ('Health 라우트',
             '로드밸런서가 컨테이너 정상 여부 확인 시 호출',
             'Flask',
             '-',
             '200 OK + 환경 요약 (secret 로드 여부 / DDB 테이블명)'),
        ]),
        ('배포 (3번 영역 : CI/CD)', [
            ('컨테이너',
             'Flask 앱을 도커 이미지로 빌드 후 ECR 푸시',
             'AWS ECS Fargate / ECR',
             'lifesync360-platform-cluster',
             'Dockerfile + buildspec.yml + taskdef.json'),
            ('파이프라인',
             'GitHub → 자동 빌드 → 무중단 배포 (Blue/Green)',
             'AWS CodePipeline + CodeBuild + CodeDeploy',
             'lifesync360-platform-pipeline',
             '실패 시 자동 롤백, 성공 시 새 버전으로 트래픽 전환'),
            ('보안 권한',
             '컨테이너가 사용하는 AWS 권한 최소화',
             'IAM Task Role',
             '-',
             'SSM 시크릿 조회 / DDB 읽기 / Lambda 호출 / Aurora 접속 만 허용'),
        ]),
    ],
}


# ── 시트 하단 API 호출 섹션 (admin V4 형식) ──────────────────────
API_SECTIONS = {
    '인증 · 로그인': ('API — 인증 · 로그인 호출', [
        ('/api/login',
         '계정 검증 통과 시 HS256 토큰 발급 후 응답에 반환',
         'app.py · api_login()',
         'POST /api/login',
         '입력: 이메일·비밀번호 / 검증: test_login_credentials.csv (5 등급 데모) / 토큰 발급: jwt.encode + SSM /lifesync360/jwt-secret'),
        ('/api/me',
         '본인 정보 조회 (이름·등급·인구통계·동의)',
         'app.py · api_me() — @require_jwt',
         'GET /api/me',
         '온프레미스(이름·동의) + DynamoDB(등급) 조합 전체'),
    ]),

    '홈 · 점수 · 캠페인': ('API — 홈 · 점수 · 캠페인 호출', [
        ('/api/dashboard',
         '내 점수 조회 (종합·건강)',
         'app.py · api_dashboard() — @require_jwt',
         'GET /api/dashboard',
         'DynamoDB lifesync_customer_result'),
        ('/api/campaigns',
         '등급별 활성 캠페인 배너 조회',
         'app.py · api_campaigns() — @require_jwt',
         'GET /api/campaigns',
         'Aurora campaign_master (등급 + 활성 + 만료 안 됨, 최대 5건)'),
    ]),

    '추천 · 상품': ('API — 추천 · 상품 호출', [
        ('/api/recommendations',
         '개인화 추천 top 10 (룰 + 캐시 + 상품 매칭)',
         'app.py · api_recommendations() — @require_jwt',
         'GET /api/recommendations',
         '6단계 흐름: DynamoDB 메타 → Redis 캐시 → 룰 매칭 → 교차 추천 → 상품 JOIN → 이력 적재. 캐시 TTL 6시간'),
        ('/product/<code>',
         '상품 상세 페이지 (SSR)',
         'app.py · product()',
         'GET /product/<product_code>',
         'Aurora product_master + 회사·카테고리·옵션'),
        ('/product/<code>/apply',
         '상품 신청 폼 페이지 (SSR)',
         'app.py · product_apply()',
         'GET /product/<product_code>/apply',
         'Aurora product_master 에서 신청용 상품 정보 추출'),
        ('/api/product/<code>/apply',
         '상품 신청 처리 (Aurora 3개 테이블 동시 갱신)',
         'app.py · api_product_apply() — @require_jwt',
         'POST /api/product/<product_code>/apply',
         '신청서 INSERT + 추천 이력 구매 플래그 UPDATE + 분석 로그 INSERT'),
    ]),

    '이벤트 · 신청 내역': ('API — 이벤트 · 신청 내역 호출', [
        ('/api/event',
         '사용자 행동 추적 (클릭·페이지뷰·신청)',
         'app.py · api_event() — @require_jwt',
         'POST /api/event',
         '8개 이벤트 종류 → 페이지 타입·클릭 종류 자동 분기 → Aurora 로그·이력 갱신'),
        ('/api/my-applications',
         '내 신청 내역 조회 (최근 50건)',
         'app.py · api_my_applications() — @require_jwt',
         'GET /api/my-applications',
         'Aurora customer_product_application + 상품·회사·카테고리 JOIN'),
    ]),

    '동의 · 설정': ('API — 동의 · 설정 호출', [
        ('/consent',
         '데이터 활용 동의 입력 화면 (SSR)',
         'app.py · consent()',
         'GET /consent',
         '8개 도메인 (은행·카드·보험·증권·다이렉트보험·헬스케어·병원·웨어러블)'),
        ('/api/consent',
         '동의 항목 저장',
         'app.py · api_consent() — @require_jwt',
         'POST /api/consent',
         '온프레미스 PrivateAPI 호출'),
        ('/settings',
         '설정 페이지 — 등급 혜택 + 동의 관리 (SSR)',
         'app.py · settings()',
         'GET /settings',
         '등급별 혜택 리스트 + 동의 항목 8종'),
    ]),

    '인프라 · 운영': ('API — 인프라 · 운영 호출', [
        ('/health',
         '컨테이너 정상 여부 확인 (ALB 헬스체크)',
         'app.py · health()',
         'GET /health',
         '응답: status / secret 로드 여부 / DDB 테이블명 / 빌드 git rev'),
    ]),
}


# ── 빌드 ──────────────────────────────────────────────────────────
def build():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name, sections in SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        for col, w in COL_WIDTHS.items():
            ws.column_dimensions[col].width = w

        col_widths_list = [COL_WIDTHS[c] for c in ['A', 'B', 'C', 'D', 'E']]

        def estimate_height(row_data):
            """텍스트 길이 + 컬럼 너비로 wrap 줄 수 추정 → 행 높이 픽셀 반환."""
            max_lines = 1
            for v, w in zip(row_data, col_widths_list):
                s = str(v) if v is not None else ''
                # 컬럼 너비 1 단위 ≈ 한글 0.5자 / ASCII 1자
                per_line_units = max(1, int(w * 1.7))
                for line in s.split('\n'):
                    units = sum(2 if ord(c) > 127 else 1 for c in line)
                    lines = max(1, (units + per_line_units - 1) // per_line_units)
                    max_lines = max(max_lines, lines)
                max_lines = max(max_lines, s.count('\n') + 1)
            # 12pt 폰트 한 줄 약 18~20px, padding 포함
            return max(22, 19 * max_lines + 8)

        def write_section(title, headers, rows, start_row):
            r = start_row
            ws.cell(r, 1, title).font = FONT_BOLD
            ws.cell(r, 1).fill = FILL_TITLE
            ws.cell(r, 1).alignment = Alignment(vertical='center', wrap_text=True)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
            ws.row_dimensions[r].height = 26
            r += 1
            for c, h in enumerate(headers, start=1):
                cell = ws.cell(r, c, h)
                cell.font = FONT_BOLD
                cell.fill = FILL_HEAD
                cell.alignment = ALIGN_C
                cell.border = BORDER
            ws.row_dimensions[r].height = 30
            r += 1
            for data in rows:
                for c, v in enumerate(data, start=1):
                    cell = ws.cell(r, c, v)
                    cell.font = FONT
                    cell.alignment = ALIGN_L
                    cell.border = BORDER
                ws.row_dimensions[r].height = estimate_height(data)
                r += 1
            return r + 1

        r = 1
        for section_title, rows in sections:
            r = write_section(section_title, HEADERS, rows, r)

        if sheet_name in API_SECTIONS:
            api_title, api_rows = API_SECTIONS[sheet_name]
            r = write_section(api_title, HEADERS_API, api_rows, r)

    wb.save(OUT)
    print(f'saved: {OUT}')


if __name__ == '__main__':
    build()
