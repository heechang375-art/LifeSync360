"""관리자 대시보드 설계서 V5 빌드 — V4 형식 + API 호출 주기/표시 방식 컬럼 추가.

사용: PYTHONIOENCODING=utf-8 python docs/build_admin_xlsx.py
출력: 관리자 대시보드 설계서 V5.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = 'C:/users/campus3S026/LS/관리자 대시보드 설계서 V5.xlsx'

# ── 스타일 (V4 동일 — 맑은 고딕 12pt, D9E1F2/FFF2CC) ─────────────
FONT       = Font(name='맑은 고딕', size=12)
FONT_BOLD  = Font(name='맑은 고딕', size=12, bold=True)
FILL_TITLE = PatternFill('solid', fgColor='D9E1F2')
FILL_HEAD  = PatternFill('solid', fgColor='FFF2CC')
ALIGN_C    = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_L    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
BORDER     = Border(*[Side(style='thin', color='BFBFBF')] * 4)

COL_WIDTHS     = {'A': 24, 'B': 38, 'C': 30, 'D': 36, 'E': 60}
COL_WIDTHS_API = {'A': 30, 'B': 36, 'C': 32, 'D': 28, 'E': 38, 'F': 16, 'G': 30}
HEADERS        = ['기능', '설명', '데이터 소스', '테이블 / 객체', '컬럼 / 비고']
HEADERS_API    = ['API', '설명', '구현 위치', '엔드포인트', '데이터 소스', '호출 주기', '표시 방식']


# ── 데이터 (각 시트 = (섹션 제목, [row, ...]) list) ─────────────
SHEETS = {
    '전체 현황 (dashboard)': [
        ('KPI Cards — 고객/추천 현황 (3×3 = 9, 화면 상단)', [
            ('통합 고객 수',          '전체 100% — 1,000,000명',                'On-Prem MySQL (ls-db)',
             'master_customer',       "COUNT(*) WHERE customer_status='ACTIVE'"),
            ('플랫폼 가입자',         '전체의 30% — 300,000명',                  'On-Prem MySQL',
             'users',                  'COUNT(DISTINCT user_id) — 가입 완료자'),
            ('분석 대상 고객',        '가입자의 20% — 60,000명',                'On-Prem MySQL',
             'users + consent_master', 'INNER JOIN consent WHERE wearable=Y AND healthcare=Y'),
            ('AI 추천 상태',          'Vertex AI · 오늘 04:30 갱신',             'DynamoDB',
             'lifesync_customer_result','MAX(update_time) + 모델 상태 표시'),
            ('누적 추천 이력',        '487,290건 (전체 누적)',                   'AWS Aurora MySQL',
             'customer_recommend_history','COUNT(*) — recommended_at IS NOT NULL'),
            ('누적 활동 로그',        '12.8M건',                                'AWS Aurora MySQL',
             'customer_dashboard_log',  'COUNT(*) — log_at 기준'),
            ('추천 CTR (클릭률)',     '14.2%',                                  'AWS Aurora MySQL',
             'customer_recommend_history',"SUM(clicked_flag='Y') / COUNT(*) * 100"),
            ('구매 전환율 (CVR)',     '9.8%',                                   'AWS Aurora MySQL',
             'customer_recommend_history',"SUM(purchased_flag='Y') / SUM(clicked_flag='Y') * 100"),
            ('Redis Cache 수',        '54,890 keys · TTL 6h',                  'AWS ElastiCache CloudWatch',
             'rec:{global_id}',        'CloudWatch CurrItems metric — Redis 캐시 키 수'),
        ]),
        ('Cloud Status — AWS / GCP / On-Prem (3카드, 가로)', [
            ('AWS 클라우드',          '7 영역 헬스 (RDS / DDB / Redis / ECS / ALB / S3 / Lambda)',
             'AWS describe APIs',     '_cloud3_from_aws()',     '_ping_cloud_status 결과 + Lambda lifesync- count + EC2 Tag=lifesync count'),
            ('GCP 클라우드',          '3 / 3 정상 (BigQuery · Vertex AI · Cloud Run)',
             'GCP Cloud Monitoring',   '_stub_gcp_status()',     '서비스별 헬스 체크 (GCP 자격증명)'),
            ('On-Premises',           '3 / 3 정상 (ls-db · ls-token · ls-api)',
             'PrivateAPI (VPN 경유)',   '_check_onprem_health()','strongSwan VPN 통한 ping'),
        ]),
        ('S3 Ingestion — 데이터 적재 현황 (5카드)', [
            ('Raw Bucket 총 파일',    'lifesync-* 다중 버킷 누적 객체 수',       'AWS CloudWatch · S3 metric',
             'AWS/S3 NumberOfObjects','24h 평균 metric — 빠른 응답 (paginate X)'),
            ('금일 적재 건수',        'KST 오늘 dt= 파티션 객체 수',             'AWS S3 list_objects_v2',
             'lifesync-raw/{domain}/dt={today}/','도메인 8 × MaxKeys 1000'),
            ('Kinesis 페이로드',      '실시간 wearable 적재 (오늘 dt)',          'AWS S3 list_objects_v2',
             'lifesync-raw/wearable/dt={today}/','MaxKeys 1000'),
            ('그룹사 적재량',         '전체 lifesync-* 버킷 총 사이즈',          'AWS CloudWatch · S3 metric',
             'AWS/S3 BucketSizeBytes', 'StandardStorage 24h 평균 (GB 단위)'),
            ('최근 업로드',           '마지막 적재 객체 시각 (KST)',             'AWS S3 list_objects_v2',
             'lifesync-raw/{domain}/dt={today}/','sorted by LastModified DESC → first'),
        ]),
        ('최근 업로드 파일 (S3 EventBridge — 5건 표)', [
            ('파일 표 5건',           '시각 / 도메인 뱃지 / 파일명 / 크기',     'AWS S3 list_objects_v2',
             'lifesync-raw/{domain}/dt={today}/','도메인 8 × MaxKeys 10 → 정렬 → 상위 5건 (KST)'),
            ('도메인 뱃지',           'BANK/CARD/INS/SEC/HLT/HOS/ONI/IOT 8종',  'app.py _BADGE_MAP',
             '-',                       '파일 prefix 첫 디렉토리 → 뱃지 매핑'),
        ]),
    ],

    '고객 통합 프로필 (users)': [
        ('고객 검색 (1번 영역, 상단)', [
            ('global_id 입력',        'G000000001 형식. 1,000,000명 모두 조회 가능',
             '관리자 UI',              'forms.search',           '입력 → /users?q={global_id} 이동 → SSR'),
            ('검색 결과 단일 프로필', '검색한 1명의 통합 정보',                 'On-Prem MySQL + DynamoDB',
             'master_customer + customer_360_profile + lifesync_customer_result','users() 핸들러에서 3 소스 JOIN'),
        ]),
        ('인구통계 / PII (2번 영역, 좌측 상단 박스)', [
            ('이름 (마스킹)',         '김*수 형태 — 풀네임 별도 PII 조회 시',  'On-Prem 토큰화 (ls-token)',
             'customer_pii_secure',    '토큰 detokenize 후 마스킹'),
            ('연령대 · 성별',         '20~60대+, M/F',                          'On-Prem MySQL',
             'customer_360_profile',   'age_band, gender'),
            ('지역 · 소득 등급',      'SEOUL/GYEONGGI 등',                      'On-Prem MySQL',
             'customer_360_profile',   'region, income_tier, asset_tier'),
            ('가입 · 마스터',         '가입일 / 최근 로그인 / 고객 상태',       'On-Prem MySQL',
             'master_customer',        'first_created_dt, last_login_dt, customer_status'),
        ]),
        ('AI 점수 (3번 영역, 4 카드)', [
            ('LifeSync 종합 점수',    'dynamic_score 0~100',                    'AWS DynamoDB',
             'lifesync_customer_result','dynamic_score (Vertex AI 일배치 산출)'),
            ('건강 점수',             'health_score 0~100',                     'AWS DynamoDB',
             'lifesync_customer_result','health_score'),
            ('금융 점수',             'fin_score 0~100',                        'AWS DynamoDB',
             'lifesync_customer_result','fin_score'),
            ('VIP 확률',              'vip_prob 0.00~1.00',                     'AWS DynamoDB',
             'lifesync_customer_result','vip_prob (ML 예측)'),
        ]),
        ('동의 현황 (4번 영역, 8 뱃지)', [
            ('8 도메인 동의',         'BANK / CARD / INSURANCE / SECURITIES / HEALTHCARE / HOSPITAL / ONLINE_INSURANCE / WEARABLE',
             'AWS S3 동의 스냅샷',     's3://lifesync-curated/consent/dt=YYYY-MM-DD/{global_id}.json','consent_snapshot_aggregator Lambda 가 매일 KST 03:00 적재'),
        ]),
        ('보유 / 추천 / 활동 (5~7번 영역, 우측)', [
            ('보유 상품 뱃지',        '현재 보유 중인 상품 카테고리',           'On-Prem MySQL',
             'customer_product_holdings','user_id 기준 active 상품'),
            ('Top-N 추천 상품',       'Redis Personalized Top 3 또는 Aurora 추천 이력',
             'AWS ElastiCache / Aurora','rec:{global_id} (Redis) / customer_recommend_history',
             'Redis 캐시 우선, miss 시 Aurora SELECT'),
            ('교차판매 추천',         'cross_sell_rule 기반 X→Y 매핑',          'AWS Aurora MySQL',
             'cross_sell_rule + customer_product_holdings','보유 상품 X → 추천 상품 Y 규칙'),
            ('최근 추천 이력',        '시각 / 상품 / 클릭/구매 여부',           'AWS Aurora MySQL',
             'customer_recommend_history JOIN product_master','ORDER BY recommended_at DESC LIMIT 20'),
            ('최근 활동 로그',        '시각 / 액션 (login/view/click/purchase)','AWS Aurora MySQL',
             'customer_dashboard_log', 'ORDER BY log_at DESC LIMIT 20'),
        ]),
    ],

    'AI 추천 (ai)': [
        ('핵심 추천 지표 (1번 영역, 4 KPI)', [
            ('추천 CTR',              '14.2% — 일정 기간 동안 노출 대비 클릭', 'AWS Lambda CloudWatch',
             'lifesync-recommendation-engine-lambda Invocations','1h sum'),
            ('구매 전환율 CVR',       '9.8% — 클릭 대비 구매',                  'AWS Lambda CloudWatch',
             'lifesync-ingest-lambda Invocations','1h sum'),
            ('AI 모델 정확도',        '0.42 vip_prob 평균',                     'AWS DynamoDB scan',
             'lifesync_customer_result','AVG(vip_prob) / signup / rec 효율 (계산)'),
            ('분석 대상 고객',        '60,000명',                                'On-Prem',
             'consent_master + users',  '동의 완료 + 가입 완료 교집합'),
        ]),
        ('7일 추천 성과 추이 (2번 영역, 막대+선 SVG)', [
            ('막대 — 추천 건수',      '일별 추천 수 7일',                       'AWS Aurora MySQL',
             'customer_recommend_history','GROUP BY DATE(recommended_at) — 최근 7일'),
            ('선 — CTR / CVR',        '일별 CTR/CVR 7일',                       'AWS Aurora MySQL',
             'customer_recommend_history','일별 SUM(clicked) / SUM(purchased) 계산'),
        ]),
        ('상품 TOP 10 (3번 영역, 우측)', [
            ('순위 / 상품 / 추천수 / CTR','추천 상위 10 상품',                  'AWS Aurora MySQL',
             'customer_recommend_history JOIN product_master','GROUP BY product_id ORDER BY COUNT DESC LIMIT 10'),
        ]),
        ('세그먼트 & 분포 (4번 영역, 3 카드)', [
            ('카테고리별 도넛',       '카테고리별 CTR 비중',                    'AWS Aurora MySQL',
             'category_master',        'GROUP BY category_code'),
            ('연령대별 추천 성과',    '20s/30s/40s/50s/60s+ × CTR/CVR',          'AWS Aurora / On-Prem JOIN',
             'customer_360_profile + customer_recommend_history','age_band 기준 GROUP BY'),
            ('고객 등급 분포',        'VIP/GOLD/SILVER/BASIC/CARE 5등급',       'AWS DynamoDB',
             'lifesync_customer_result','GROUP BY dynamic_grade'),
        ]),
        ('BigQuery 분석 (5번 영역, 3 카드)', [
            ('Feature 분포',          'Vertex AI feature_importance',           'GCP BigQuery / GCP Vertex AI',
             'recommendation_mart',    'AI 입력 feature 의 중요도 분포'),
            ('추천 데이터 분석',      'action_code 별 추천 수',                 'GCP BigQuery',
             'recommendation_mart',    'GROUP BY action_code — 전 7일'),
            ('고객 인사이트',         'BigQuery ad-hoc 분석 KPI',               'GCP BigQuery',
             'customer_summary_mart',  'BQ 쿼리 결과 (label/value/sub)'),
        ]),
        ('AI 모델 평가 (6번 영역, 2 카드)', [
            ('AI 예측 출현 분포',     'dynamic_score 5 bucket 히스토그램',      'AWS DynamoDB scan',
             'lifesync_customer_result','dynamic_score 0-20/20-40/40-60/60-80/80-100'),
            ('Precision / Recall',    'VIP / Churn / NBA 모델 평가',            'AWS Aurora MySQL',
             'model_performance_history','정밀도·재현율 + 재훈련 알람'),
        ]),
    ],

    'Network & Wearable (ops)': [
        ('인프라 토폴로지 (1번 영역, 상단 다이어그램)', [
            ('AWS / GCP / On-Prem',   '3 클라우드 토폴로지',                    'AWS describe + GCP + On-Prem 메타데이터',
             'TGW + VPN + GCP Interconnect','TGW · VPN · GCP Interconnect 라인 표시'),
        ]),
        ('AWS 3 VPC (2번 영역, 3 카드)', [
            ('Platform VPC',          'Aurora · Redis · DynamoDB · Lambda',     'AWS describe APIs',
             'vpc-* + Aurora + Redis + DDB + Lambda','각 서비스 상태 + 인스턴스 카운트'),
            ('Data VPC',              'S3 · Kinesis · Glue · EMR',              'AWS describe APIs',
             'vpc-* + S3 + Kinesis + Glue + EMR','Kinesis 스트림 상태 + Glue 최근 run'),
            ('GroupVM VPC',           'EC2 7개 (BANK/CARD/SEC/INS/ONI/HLT/HOS)','AWS EC2',
             'EC2 Tag=계열사 7 인스턴스', '계열사별 EC2 인스턴스 상태'),
        ]),
        ('Connectivity / GCP / On-Prem (3번 영역, 3 카드)', [
            ('AWS Connectivity',      'TGW · VPN · VPC Peering',                'AWS EC2 describe',
             'tgw-* / vpn-* / vpc-peering','tgw / vpn / vpc-peering 상태'),
            ('GCP',                   'BigQuery · Vertex AI · Cloud Run',       'GCP Cloud Monitoring',
             'BigQuery + Vertex AI + Cloud Run','GCP 자격증명으로 헬스 체크'),
            ('On-Prem',               'ls-db · ls-token · ls-api 헬스',         'PrivateAPI (VPN 경유)',
             'ls-db / ls-token / ls-api','strongSwan VPN 통한 ping'),
        ]),
        ('Wearable 실시간 (4번 영역, KPI 4 + 표 2)', [
            ('활성 디바이스',         '최근 3초 송신 디바이스 수',              'AWS Kinesis Stream + DynamoDB',
             'wearable_latest{global_id}','Kinesis consumer 가 매 batch 100 record DDB 갱신'),
            ('송신율',                '활성 / 등록 %',                          'AWS Kinesis Stream',
             'IncomingRecords (5분 윈도우)','CloudWatch metric 활용 — active / registered'),
            ('건강 RED 카운트',       'AHA/WHO 임상 임계 위반 (즉시 조치)',     'AWS DynamoDB anomaly_event',
             'anomaly_event scan (today)','HR<50 / HR≥120 / SpO2<90 / stress≥76'),
            ('건강 YELLOW 카운트',    '경계 영역 (모니터링)',                   'AWS DynamoDB anomaly_event',
             'anomaly_event scan (today)','HR 50-59 또는 101-119 / SpO2 90-94 / stress 51-75'),
            ('🔴 건강 RED 표',         '최근 N 명 — 시각/gid/이름/심박/SpO2/스트레스/사유','AWS DynamoDB anomaly_event',
             'anomaly_event (PK: global_id, SK: event_time)','시계열 알람 로그 — 시간순 정렬, 화면 상위 10개'),
            ('🟡 건강 YELLOW 표',     '경계 영역 N 명 동일 컬럼',               'AWS DynamoDB anomaly_event',
             'anomaly_event filter by severity','동일 (severity=YELLOW)'),
        ]),
        ('Network API 엔드포인트 (5번 영역, 하단 list)', [
            ('Network 라우트 표',     'TGW / VPN / VM / Kinesis / EMR / Local Lab','admin 라우트 메타',
             'Network API 9 엔드포인트','9 라우트 listing — name / method / desc'),
        ]),
    ],
}


# ── API 호출 표 (각 시트 하단) — V5 신설: 호출 주기 + 표시 방식 컬럼 ───
API_SECTIONS = {
    '전체 현황 (dashboard)': [
        ('GET /api/dashboard/summary', 'KPI 9 카드 집계',
         'admin-platform/app.py · _stub_aurora_summary()',
         '/api/dashboard/summary',
         'On-Prem users + Aurora 추천이력/행동로그 + DynamoDB',
         'JS 폴링 60s', 'KPI 9 카드 textContent 교체 + ts-kpi 갱신'),
        ('GET /api/dashboard/cloud3', 'AWS/GCP/On-Prem 3 카드',
         'admin-platform/app.py · _cloud3_from_aws()',
         '/api/dashboard/cloud3',
         'RDS/DDB/EC/ECS/ALB/S3 describe + Lambda list + EC2 describe',
         'JS 폴링 60s', 'Cloud 3 카드 state/sub textContent + ts-cloud'),
        ('GET /api/s3/status', 'S3 INGESTION 5 카드',
         'admin-platform/app.py · _s3_status_cards()',
         '/api/s3/status',
         'CloudWatch S3 metric (NumberOfObjects + BucketSizeBytes) + 오늘 dt KeyCount',
         'JS 폴링 60s', 'S3 5 카드 value/note textContent + ts-s3'),
        ('GET /api/dashboard/uploads', '최근 업로드 5건 표',
         'admin-platform/app.py · _uploads_from_s3(limit=5)',
         '/api/dashboard/uploads',
         'S3 list_objects_v2 도메인 8 × MaxKeys 10 → sort → 5 (KST)',
         'JS 폴링 60s', '<tbody id="uploads-tbody"> innerHTML 교체 + ts-uploads'),
    ],
    '고객 통합 프로필 (users)': [
        ('GET /api/customer/profile/<global_id>', '고객 통합 프로필 (PII + 마스터 + 360_profile)',
         'admin-platform/app.py · api_customer_profile',
         '/api/customer/profile/{gid}',
         'On-Prem PrivateAPI get_profile + S3 consent snapshot',
         'SSR / 수동', '페이지 로드 시 1회 — 검색 후 페이지 이동'),
        ('GET /api/customer/ai-result/<global_id>', 'AI 점수 (4종)',
         'admin-platform/app.py · _stub_ai_result',
         '/api/customer/ai-result/{gid}',
         'DynamoDB get_item (HASH+RANGE 키 필요)',
         'SSR / 수동', '페이지 로드 시 1회'),
        ('GET /api/customer/recommend/<global_id>', 'Personalized Top-N 추천',
         'admin-platform/app.py · _stub_redis_personalized',
         '/api/customer/recommend/{gid}',
         'Redis GET rec:{global_id} → Aurora fallback',
         'SSR / 수동', '페이지 로드 시 1회'),
        ('GET /api/customer/history/<global_id>', '추천 이력 (최근 N)',
         'admin-platform/app.py · _stub_aurora_history',
         '/api/customer/history/{gid}',
         'Aurora customer_recommend_history JOIN product_master',
         'SSR / 수동', '페이지 로드 시 1회'),
        ('GET /api/customer/activity/<global_id>', '활동 로그 (최근 N)',
         'admin-platform/app.py · _stub_aurora_activity',
         '/api/customer/activity/{gid}',
         'Aurora customer_dashboard_log',
         'SSR / 수동', '페이지 로드 시 1회'),
    ],
    'AI 추천 (ai)': [
        ('GET /api/ai/kpi4', '상단 4 KPI',
         'admin-platform/app.py · _ai_kpi4_from_aws()',
         '/api/ai/kpi4',
         'Lambda CloudWatch Invocations (recommendation / ingest 1h)',
         'JS 폴링 300s', 'KPI 4 카드 value/sub + ts-ai-kpi'),
        ('GET /api/ai/summary', 'Vertex AI 메트릭 + DynamoDB 분포',
         'admin-platform/app.py · _stub_ai_summary',
         '/api/ai/summary',
         'Vertex AI stub + DDB scan (dynamic_score 분포)',
         '폴링 미적용 (SSR)', '페이지 로드 시 1회 — Jinja2 SVG 좌표 계산'),
        ('GET /api/ai/recommend-stats', '7일 추이 / 카테고리 / TOP10',
         'admin-platform/app.py · _stub_recommend_stats',
         '/api/ai/recommend-stats',
         'Aurora SUM(clicked) / SUM(purchased) GROUP BY DATE',
         '폴링 미적용 (SSR)', '페이지 로드 시 1회 — SVG 차트 정적'),
        ('GET /api/bigquery/analytics', 'BigQuery ad-hoc (recommendation_mart / customer_summary / prediction_result)',
         'admin-platform/app.py · _stub_bigquery_analytics(kind=)',
         '/api/bigquery/analytics?kind=...',
         'GCP BigQuery (자격증명 필요)',
         '수동 호출', 'API 테스트용 — 화면 자동 호출 X'),
    ],
    'Network & Wearable (ops)': [
        ('GET /stream/wearable', 'Wearable 실시간 — KPI 4 + RED/YELLOW 표',
         'admin-platform/app.py · stream_wearable()',
         '/stream/wearable',
         'AWS Kinesis Stream consumer + DynamoDB anomaly_event',
         'SSE 3s push', '<tbody> innerHTML 교체 + KPI textContent + ts-we-* 3건'),
        ('GET /api/ops/wearable', 'SSE 폴백 (단발 JSON)',
         'admin-platform/app.py · api_ops_wearable()',
         '/api/ops/wearable',
         'AWS DynamoDB wearable_latest + anomaly_event scan',
         '수동 호출', 'SSE 비대응 클라이언트용 — 화면 자동 호출 X'),
        ('GET /api/network/tgw', 'Transit Gateway 상태',
         'admin-platform/app.py · api_network_tgw',
         '/api/network/tgw',
         'EC2 describe_transit_gateways + attachments',
         '수동 호출', 'API 테스트용 (ops 화면 폴링 미연결 — 라우트 보유)'),
        ('GET /api/network/vpn', 'VPN 터널 상태',
         'admin-platform/app.py · api_network_vpn',
         '/api/network/vpn',
         'EC2 describe_vpn_connections + VgwTelemetry',
         '수동 호출', 'API 테스트용'),
        ('GET /api/vm/group', 'Group VM EC2 (7 계열사)',
         'admin-platform/app.py · api_vm_group',
         '/api/vm/group',
         'EC2 describe_instances (tag Project=lifesync)',
         '수동 호출', 'API 테스트용'),
        ('GET /api/vm/wearable', 'Wearable VM + Custom metric',
         'admin-platform/app.py · api_vm_wearable',
         '/api/vm/wearable',
         'EC2 describe + CloudWatch LifeSync/Wearable namespace',
         '수동 호출', 'API 테스트용'),
        ('GET /api/kinesis/status', 'Kinesis 스트림 상태',
         'admin-platform/app.py · api_kinesis_status',
         '/api/kinesis/status',
         'Kinesis describe_stream',
         '수동 호출', 'API 테스트용'),
        ('GET /api/emr/status', 'EMR 클러스터 상태',
         'admin-platform/app.py · api_emr_status',
         '/api/emr/status',
         'EMR list_clusters',
         '수동 호출', 'API 테스트용 (현재 EMR 미배포 — 빈 응답)'),
        ('GET /api/admin/local-lab-status', 'On-Prem Local Lab 헬스',
         'admin-platform/app.py · api_admin_local_lab_status',
         '/api/admin/local-lab-status',
         'PrivateAPI onprem-query local_lab_status action',
         '수동 호출', 'API 테스트용'),
    ],
}


# ── 행 높이 자동 계산 (한글 2 / ASCII 1 가중치) ─────────────────
def estimate_height(row_data, widths):
    max_lines = 1
    for v, w in zip(row_data, widths):
        s = str(v) if v is not None else ''
        per_line_units = max(1, int(w * 1.7))
        for line in s.split('\n'):
            units = sum(2 if ord(c) > 127 else 1 for c in line)
            lines = max(1, (units + per_line_units - 1) // per_line_units)
            max_lines = max(max_lines, lines)
        max_lines = max(max_lines, s.count('\n') + 1)
    return max(22, 19 * max_lines + 8)


# ── 시트 빌드 ───────────────────────────────────────────────────
def write_section(ws, row, title, headers, rows, col_widths_letters):
    n_cols = len(headers)
    widths = [col_widths_letters[chr(ord('A') + i)] for i in range(n_cols)]

    # 섹션 제목 (병합)
    ws.cell(row, 1, title).fill = FILL_TITLE
    ws.cell(row, 1).font  = FONT_BOLD
    ws.cell(row, 1).alignment = ALIGN_L
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    ws.row_dimensions[row].height = 26
    row += 1

    # 컬럼 헤더
    for i, h in enumerate(headers):
        c = ws.cell(row, i + 1, h)
        c.fill, c.font, c.alignment, c.border = FILL_HEAD, FONT_BOLD, ALIGN_C, BORDER
    ws.row_dimensions[row].height = 24
    row += 1

    # 데이터 행
    for r in rows:
        for i, v in enumerate(r):
            c = ws.cell(row, i + 1, v)
            c.font, c.alignment, c.border = FONT, ALIGN_L, BORDER
        ws.row_dimensions[row].height = estimate_height(r, widths)
        row += 1

    return row + 1   # 섹션 사이 빈 줄


def build():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name, sections in SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        # 컬럼 폭
        for letter, w in COL_WIDTHS.items():
            ws.column_dimensions[letter].width = w

        row = 1
        for title, rows in sections:
            row = write_section(ws, row, title, HEADERS, rows, COL_WIDTHS)

        # API 호출 표 (7컬럼)
        api_rows = API_SECTIONS.get(sheet_name, [])
        if api_rows:
            for letter, w in COL_WIDTHS_API.items():
                ws.column_dimensions[letter].width = w
            row = write_section(ws, row, '🌐 API — 화면 데이터 호출 (V5 신설: 호출 주기 + 표시 방식)',
                                HEADERS_API, api_rows, COL_WIDTHS_API)

    wb.save(OUT)
    print(f'saved: {OUT}')


if __name__ == '__main__':
    build()
